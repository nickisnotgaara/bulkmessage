"""Contacts loading from Excel."""

from __future__ import annotations

from typing import Any

from openpyxl import load_workbook

from .wappi import normalize_phone


def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    return str(h).strip().lower().replace("ё", "е")


def load_contacts(path: str) -> list[dict]:
    """Читает Excel и возвращает [{phone, name, category}].

    Поддерживает разные заголовки:
      - phone: телефон / номера телефонов / номер телефона / добавочный номер / phone
        (пробует каждый подходящий столбец по очереди, пока не найдёт валидный номер)
      - name: имя контакта / имя / фио / name
        (наименование контакта НЕ используется как имя — там длинное описание)
      - category: категория / статус / category / status
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_norm_header(h) for h in rows[0]]

    phone_keys = (
        "номера телефонов",
        "номер телефона",
        "телефон",
        "добавочный номер",
        "phone",
    )
    # Имя — СТРОГО из столбца «Имя контакта». Любой другой «имя*» игнорируем.
    name_keys_exact = ("имя контакта",)
    cat_keys = ("категория", "статус", "category", "status")

    def find_idx(keys: tuple[str, ...]) -> list[int]:
        """Точные совпадения заголовков, в порядке переданных ключей."""
        result: list[int] = []
        for k in keys:
            for i, h in enumerate(headers):
                if h == k and i not in result:
                    result.append(i)
        return result

    phone_idxs = find_idx(phone_keys)
    name_idxs = find_idx(name_keys_exact)
    cat_idxs = find_idx(cat_keys)

    def pick_phone(row) -> str:
        for i in phone_idxs:
            if i < len(row) and row[i]:
                p = normalize_phone(row[i])
                if p:
                    return p
        return ""

    def pick_name(row) -> str:
        """Берём имя только из столбцов «Имя контакта» / «Имя» / «ФИО».

        Если ничего не нашли — возвращаем пустую строку. Никогда
        не используем «Наименование контакта» (там длинные описания,
        и для категории риэлторов templates.py подставит «коллега»).
        """
        for i in name_idxs:
            if i < len(row) and row[i]:
                v = str(row[i]).strip()
                if v:
                    return v
        return ""

    def pick_category(row) -> str:
        for i in cat_idxs:
            if i < len(row) and row[i]:
                v = str(row[i]).strip()
                if v:
                    return v
        return ""

    contacts: list[dict] = []
    seen: set[str] = set()
    skipped_no_phone = 0
    for row in rows[1:]:
        if not row:
            continue
        phone = pick_phone(row)
        if not phone:
            skipped_no_phone += 1
            continue
        if phone in seen:
            continue
        name = pick_name(row)
        category = pick_category(row)
        seen.add(phone)
        contacts.append({"phone": phone, "name": name, "category": category})
    return contacts
